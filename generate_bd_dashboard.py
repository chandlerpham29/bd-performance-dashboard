#!/usr/bin/env python3
"""Generate BD Performance Dashboard scoping workbook and HTML demo."""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

WORKSPACE = "/Users/elsa/Documents/Cursor Project /Sales Performance Dashboard "
TEMPLATE = os.path.join(WORKSPACE, "B2B QBR Dashboard_Scoping (Template) .xlsx")
OUTPUT_XLSX = os.path.join(WORKSPACE, "BD Performance Dashboard_Scoping.xlsx")
OUTPUT_HTML = os.path.join(WORKSPACE, "bd-performance-dashboard-demo.html")

HEADER_FILL = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
HEADER_FONT = Font(bold=True, color="000000")
WRAP = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, col_count):
    for c in range(1, col_count + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP


def write_row(ws, row_idx, values):
    for c, val in enumerate(values, start=1):
        cell = ws.cell(row_idx, c, val)
        cell.alignment = WRAP


def build_scoping_rows():
    return [
        {
            "id": "BDPERF-1",
            "title": "Scorecard — Target Attainment",
            "objective": "Show quota target, actual bookings, new business pipeline created, activity targets, and attainment for the selected period so BD reps and leadership can monitor pacing in real time and quickly spot where performance is on or off track.",
            "chart": "Table",
            "purpose": "Single view of quota and activity target attainment for weekly BD review.",
            "priority": "P0",
            "effort": 2,
            "alignment": "Revenue / Quota",
            "audience": "BD Reps, BD Leadership",
            "status": "New",
        },
        {
            "id": "BDPERF-2",
            "title": "# of Won Deals Breakdown by Pipeline",
            "objective": "Show closed-won deal count and ACV by pipeline type for the selected period so reps can see which funnels are driving wins and where to focus follow-up.",
            "chart": "Table",
            "purpose": "Show which pipelines drive closed-won volume for the rep in the selected period.",
            "priority": "P0",
            "effort": 1,
            "alignment": "Pipeline Progress",
            "audience": "BD Reps, BD Leadership",
            "status": "New",
        },
        {
            "id": "BDPERF-3",
            "title": "# of Won Deals MoM",
            "objective": "Show month-over-month closed-won deal volume so reps can track won-deal momentum, spot improving or declining trends, and adjust activity accordingly.",
            "chart": "Bar Chart",
            "purpose": "Track won-deal momentum over time at the grain matching the selected filter.",
            "priority": "P0",
            "effort": 1.5,
            "alignment": "Pipeline Progress",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-4",
            "title": "Open Pipeline and Coverage",
            "objective": "Show open pipeline amount, deal count, and coverage vs remaining quota gap so reps can assess whether forecasted pipeline is sufficient to close the period and prioritize deals by forecast category.",
            "chart": "Table",
            "purpose": "Summarize open pipeline health and whether forecasted pipeline covers remaining quota gap.",
            "priority": "P0",
            "effort": 2,
            "alignment": "Pipeline Progress",
            "audience": "BD Reps, BD Leadership",
            "status": "New",
        },
        {
            "id": "BDPERF-5",
            "title": "Open Pipeline MoM",
            "objective": "Show monthly open pipeline ACV trend split by pipeline type so reps can understand how total open pipe and pipeline mix evolve over time and whether new business vs renewal pipeline is building as expected.",
            "chart": "Stacked Bar Chart",
            "purpose": "Alternative view of open pipeline amount in open stages, showing pipeline mix over time.",
            "priority": "P0",
            "effort": 2,
            "alignment": "Pipeline Progress",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-6",
            "title": "# of New Biz Deal Created",
            "objective": "Show new business deals created vs target for the selected period so reps can monitor pipeline creation pacing and flag underperformance before it impacts future bookings.",
            "chart": "KPI Number + MoM Bar Chart",
            "purpose": "Monitor new pipeline creation volume against pacing target.",
            "priority": "P0",
            "effort": 2,
            "alignment": "Pipeline Creation",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-7",
            "title": "# of New Demo Completed",
            "objective": "Show demos completed vs target for the selected period so reps can track mid-funnel conversion activity, compare to pacing goals, and identify coaching opportunities.",
            "chart": "KPI Number + MoM Bar Chart",
            "purpose": "Track demo completion activity against target pacing.",
            "priority": "P0",
            "effort": 2,
            "alignment": "Pipeline Progress",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-8",
            "title": "# of New Trial Completed",
            "objective": "Show trials completed vs target for the selected period so reps can measure late-stage evaluation activity against expectations and adjust deal strategy where trials stall.",
            "chart": "KPI Number + MoM Bar Chart",
            "purpose": "Track trial completion activity against target pacing.",
            "priority": "P0",
            "effort": 2,
            "alignment": "Pipeline Progress",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-9",
            "title": "Deals Progressed in Stage",
            "objective": "List deals that advanced stage in the selected period so reps and managers can surface momentum, prioritize follow-up, and coach on next steps for active opportunities.",
            "chart": "Table",
            "purpose": "Surface deals with recent momentum for coaching and follow-up.",
            "priority": "P1",
            "effort": 2,
            "alignment": "Pipeline Progress",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-10",
            "title": "Meetings with Existing Clients",
            "objective": "Show meeting volume with existing clients for the selected period so reps can measure account engagement effort and its relationship to expansion pipeline.",
            "chart": "KPI Number",
            "purpose": "Measure engagement on active accounts and expansion pipeline.",
            "priority": "P0",
            "effort": 1,
            "alignment": "Activity / Engagement",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-11",
            "title": "Meetings with New Leads",
            "objective": "Show first meetings with new leads for the selected period so reps can track top-of-funnel meeting generation and whether prospecting activity is converting to conversations.",
            "chart": "KPI Number",
            "purpose": "Track top-of-funnel meeting generation for new business.",
            "priority": "P0",
            "effort": 1,
            "alignment": "Activity / Engagement",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-12",
            "title": "Assigned Leads",
            "objective": "Show inbound leads assigned to the rep for the selected period so reps can understand lead flow volume and balance workload against outbound effort.",
            "chart": "KPI Number",
            "purpose": "Show inbound/assigned lead flow to the rep.",
            "priority": "P1",
            "effort": 1,
            "alignment": "Lead Generation",
            "audience": "BD Reps",
            "status": "New",
        },
        {
            "id": "BDPERF-13",
            "title": "Emails Sent",
            "objective": "Show outbound email volume for the selected period so reps can track prospecting communication effort alongside meetings and pipeline outcomes.",
            "chart": "KPI Number",
            "purpose": "Track outbound communication effort.",
            "priority": "P0",
            "effort": 1,
            "alignment": "Activity / Prospecting",
            "audience": "BD Reps",
            "status": "New",
        },
    ]


def build_metrics_rows():
    rows = []

    filters = [
        ("Close Period", "Refers to the time period used to scope all dashboard metrics (Quarter, Month, Week, or Year), aligned to deal close date or activity date as applicable.", "Dropdown: Quarter / Month / Week / Year", "Close Date (derived period)", "Yes", "Quarter — Current Q", "n/a", "Interact with ALL dashboard metrics.", "BD Reps, BD Leadership"),
        ("BD Rep (Owner)", "Refers to the HubSpot deal owner / BD representative responsible for the deal or activity being measured.", "Dropdown of active BD owners", "Deal Owner, hubspot_owner_id", "Yes", "Current User (default)", "n/a", "Primary rep-level filter; all metrics scoped to selected owner.", "BD Reps, BD Leadership"),
    ]
    for name, definition, pseudo, fields, use_filter, default, threshold, interactions, audience in filters:
        rows.append({
            "request_id": "Top Filters",
            "metric_name": name,
            "tab": "Global Filters",
            "definition": definition,
            "sql": "n/a",
            "window": "n/a",
            "fmt": "Dropdown / Date",
            "agg": "n/a",
            "threshold": threshold,
            "fields": fields,
            "std": "n/a",
            "filter": use_filter,
            "default_filter": default,
            "chart": "n/a",
            "purpose": "Scope all BD performance metrics to rep and time period.",
            "interactions": interactions,
            "audience": audience,
            "status": "New",
        })

    metrics = [
        # Scorecard
        ("BDPERF-1", "Scorecard - ACV Booked", "Scorecard", "Annual Contract Value of deals Closed Won (and Booked) during a Respective Period", 'Sum ("Deal" Amount (USD)) where Deal Stage = "Closed Won" and Close Date = Respective Period and Deal Owner = Selected BD Rep', "Selected close period", "Currency, USD", "Owner, Quarter, Month", "Attainment Green ≥80%; Amber 60–79%; Red <60%", "Deal Stage, Amount (USD), Close Date, Owner", "USD K", "No", "N/A", "Table", "Actual bookings against quota.", "Click to see deal list", "BD Reps", "New", "Traffic light on attainment"),
        ("BDPERF-1", "Scorecard - New Business Pipeline Created", "Scorecard", "Total value of new business sales opportunities (pipeline) created within the Respective Period", 'Sum ("Deal" Amount (USD)) where Create Date = Respective Period and Deal Pipeline = New Business and Deal Owner = Selected BD Rep', "Selected close period", "Currency, USD", "Owner, Month", "Attainment Green ≥80%; Amber 60–79%; Red <60%", "Create Date, Amount (USD), Pipeline, Owner", "USD K", "No", "N/A", "Table", "New business pipeline generation against target.", "Click to drill", "BD Reps", "New", "Target scales: Q=3× monthly, Week=ceil(monthly/4), Year=6×"),
        ("BDPERF-1", "Scorecard - New Business Deals Created", "Scorecard", "Number of new business deals created during the Respective Period", 'Count ("Deal") where Create Date = Respective Period and Deal Pipeline = New Business and Deal Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Month", "Green ≥ target; Red < target", "Create Date, Pipeline, Owner", "0", "No", "N/A", "Table", "Activity target attainment.", "Click to drill", "BD Reps", "New", "Target scales: Q=3× monthly, Week=ceil(monthly/4), Year=6×"),
        ("BDPERF-1", "Scorecard - # of New Demo Completed", "Scorecard", "Number of new demos that have been completed within the respective period", 'Count ("Deal") where Demo Completed Date = Respective Period and Deal Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Month", "Green ≥ target; Red < target", "Deal Stage, demo date, Owner", "0", "No", "N/A", "Table", "Demo activity attainment.", "Click to drill", "BD Reps", "New", ""),
        ("BDPERF-1", "Scorecard - # of New Trial Completed", "Scorecard", "Number of new trials that have been completed within the respective period", 'Count ("Deal") where Trial Completed Date = Respective Period and Deal Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Month", "Green ≥ target; Red < target", "Deal Stage, trial date, Owner", "0", "No", "N/A", "Table", "Trial activity attainment.", "Click to drill", "BD Reps", "New", ""),
        # Won Deals
        ("BDPERF-2", "Won Deals Breakdown - # of Deals by Pipeline", "Pipeline Progress", "Number of deals Closed Won during the Respective Period, grouped by Deal Pipeline", 'Count ("Deal") where Deal Stage = "Closed Won" and Close Date = Respective Period and Deal Owner = Selected BD Rep Group By Deal Pipeline', "Selected close period", "Count", "Owner, Pipeline", "n/a", "Deal Stage, Close Date, Pipeline, Owner", "0", "No", "N/A", "Table", "Pipeline mix of won deals.", "Click pipeline cell for deal list", "BD Reps", "New", ""),
        ("BDPERF-2", "Won Deals Breakdown - Total ACV by Pipeline", "Pipeline Progress", "Total Annual Contract Value of deals Closed Won during the Respective Period, grouped by Deal Pipeline", 'Sum ("Deal" Amount (USD)) where Deal Stage = "Closed Won" and Close Date = Respective Period and Deal Owner = Selected BD Rep Group By Deal Pipeline', "Selected close period", "Currency, USD", "Owner, Pipeline", "n/a", "Amount (USD), Close Date, Pipeline", "USD K", "No", "N/A", "Table", "Revenue mix of won deals.", "Click amount for deal list", "BD Reps", "New", ""),
        ("BDPERF-3", "Won Deals MoM - # of Deals", "Pipeline Progress", "Month-over-month count of deals Closed Won during months within the Respective Period", 'Count ("Deal") where Deal Stage = "Closed Won" and Close Date = Respective Month and Deal Owner = Selected BD Rep Group By Close Month', "Months in selected period", "Count", "Owner, Month", "n/a", "Close Date, Deal Stage, Owner", "0", "No", "N/A", "Bar Chart", "Won deal volume trend.", "Hover for monthly values", "BD Reps", "New", ""),
        # Open Pipeline
        ("BDPERF-4", "Open Pipeline and Coverage - Total Amount", "Pipeline Progress", "Total Annual Contract Value of open deals at the end of the Respective Period snapshot", 'Sum ("Deal" Amount (USD)) where Deal Stage <> "Closed Won" and Deal Stage <> "Closed Lost" and Deal Owner = Selected BD Rep', "End-of-period snapshot", "Currency, USD", "Owner, Forecast Category", "n/a", "Amount (USD), dealstage, hs_manual_forecast_category", "USD K", "No", "N/A", "Table", "Total open pipeline value.", "Click for deal list", "BD Reps", "New", "Aligns with stacked MoM bar total"),
        ("BDPERF-4", "Open Pipeline and Coverage - # of Deals", "Pipeline Progress", "Number of open deals at the end of the Respective Period snapshot, by forecast category", 'Count ("Deal") where Deal is Open and Deal Owner = Selected BD Rep Group By Forecast Category', "End-of-period snapshot", "Count", "Owner, Forecast Category", "n/a", "Deal ID, dealstage, hs_manual_forecast_category", "0", "No", "N/A", "Table", "Deal volume in pipeline.", "n/a", "BD Reps", "New", ""),
        ("BDPERF-4", "Open Pipeline and Coverage - Coverage", "Pipeline Progress", "Ratio of total open pipeline ACV to remaining quota gap, indicating pipeline sufficiency against target", "Total Open Pipeline Amount / (Quota Target − ACV Booked) for Respective Period", "End-of-period snapshot", "Number (x)", "Owner", "Green ≥3x; Amber 2–3x; Red <2x", "Open Pipeline, Gap to Target", "0.0", "No", "N/A", "Table", "Pipeline sufficiency vs quota.", "Click for open deals", "BD Reps", "New", "Traffic light on coverage"),
        ("BDPERF-4", "Open Pipeline and Coverage - Commit", "Pipeline Progress", "Annual Contract Value of open deals in the Commit forecast category at end of period", 'Sum ("Deal" Amount (USD)) where Forecast Category = "Commit" and Deal is Open and Deal Owner = Selected BD Rep', "End-of-period snapshot", "Currency, USD", "Owner, Forecast Category", "n/a", "Amount (USD), hs_manual_forecast_category", "USD K", "No", "N/A", "Table", "High-confidence pipeline.", "Drill to deals", "BD Reps", "New", ""),
        ("BDPERF-4", "Open Pipeline and Coverage - Best Case", "Pipeline Progress", "Annual Contract Value of open deals in the Best Case forecast category at end of period", 'Sum ("Deal" Amount (USD)) where Forecast Category = "Best Case" and Deal is Open and Deal Owner = Selected BD Rep', "End-of-period snapshot", "Currency, USD", "Owner, Forecast Category", "n/a", "Amount (USD), hs_manual_forecast_category", "USD K", "No", "N/A", "Table", "Moderate-confidence pipeline.", "Drill to deals", "BD Reps", "New", ""),
        ("BDPERF-4", "Open Pipeline and Coverage - Pipeline", "Pipeline Progress", "Annual Contract Value of open deals in the Pipeline forecast category at end of period", 'Sum ("Deal" Amount (USD)) where Forecast Category = "Pipeline" and Deal is Open and Deal Owner = Selected BD Rep', "End-of-period snapshot", "Currency, USD", "Owner, Forecast Category", "n/a", "Amount (USD), hs_manual_forecast_category", "USD K", "No", "N/A", "Table", "Early-stage pipeline.", "Drill to deals", "BD Reps", "New", ""),
        ("BDPERF-5", "Open Pipeline MoM - New Business", "Pipeline Progress", "Monthly open pipeline ACV in the New Business pipeline at each month-end snapshot within the Respective Period", 'Sum ("Deal" Amount (USD)) where Deal Pipeline = New Business and Deal is Open and Snapshot Date = Respective Month and Deal Owner = Selected BD Rep', "Months in selected period", "Currency, USD", "Owner, Month, Pipeline", "n/a", "Amount (USD), pipeline, snapshot date", "USD K", "No", "N/A", "Stacked Bar Chart", "New business open pipe over time.", "Stack total = table total", "BD Reps", "New", ""),
        ("BDPERF-5", "Open Pipeline MoM - Renewal Pipeline", "Pipeline Progress", "Monthly open pipeline ACV in the Renewal/Expansion pipeline at each month-end snapshot within the Respective Period", 'Sum ("Deal" Amount (USD)) where Deal Pipeline IN (Renewal, Expansion) and Deal is Open and Snapshot Date = Respective Month and Deal Owner = Selected BD Rep', "Months in selected period", "Currency, USD", "Owner, Month, Pipeline", "n/a", "Amount (USD), pipeline, snapshot date", "USD K", "No", "N/A", "Stacked Bar Chart", "Renewal open pipe over time.", "Stack total = table total", "BD Reps", "New", ""),
        # New Biz / Demo / Trial
        ("BDPERF-6", "New Biz Deals Created - Period Count", "Pipeline Progress", "Number of new business deals created during the Respective Period", 'Count ("Deal") where Create Date = Respective Period and Deal Pipeline = New Business and Deal Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Month, Week", "Green ≥ scaled target; Red < target", "Create Date, Pipeline, Owner", "0", "No", "N/A", "KPI Number", "Period total vs target.", "Click for deal list", "BD Reps", "New", "Target: month=monthly, Q=3×, week=ceil(m/4), year=6×"),
        ("BDPERF-6", "New Biz Deals Created - MoM Count", "Pipeline Progress", "Month-over-month count of new business deals created within the Respective Period", 'Count ("Deal") where Create Date = Respective Month and Deal Pipeline = New Business and Deal Owner = Selected BD Rep Group By Create Month', "Months in selected period", "Count", "Owner, Month", "Target line = monthly target", "Create Date, Pipeline, Owner", "0", "No", "N/A", "Bar Chart", "New biz creation trend.", "Hover monthly values", "BD Reps", "New", ""),
        ("BDPERF-7", "New Demo Completed - Period Count", "Pipeline Progress", "Number of demos completed during the Respective Period", 'Count ("Deal") where Demo Completed Date = Respective Period and Deal Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Month, Week", "Green ≥ scaled target; Red < target", "Deal Stage, demo date, Owner", "0", "No", "N/A", "KPI Number", "Period demo total vs target.", "Click for deal list", "BD Reps", "New", ""),
        ("BDPERF-7", "New Demo Completed - MoM Count", "Pipeline Progress", "Month-over-month count of demos completed within the Respective Period", 'Count ("Deal") where Demo Completed Date = Respective Month and Deal Owner = Selected BD Rep Group By Demo Completed Month', "Months in selected period", "Count", "Owner, Month", "Target line = monthly target", "Deal Stage, demo date, Owner", "0", "No", "N/A", "Bar Chart", "Demo completion trend.", "Hover monthly values", "BD Reps", "New", ""),
        ("BDPERF-8", "New Trial Completed - Period Count", "Pipeline Progress", "Number of trials completed during the Respective Period", 'Count ("Deal") where Trial Completed Date = Respective Period and Deal Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Month, Week", "Green ≥ scaled target; Red < target", "Deal Stage, trial date, Owner", "0", "No", "N/A", "KPI Number", "Period trial total vs target.", "Click for deal list", "BD Reps", "New", ""),
        ("BDPERF-8", "New Trial Completed - MoM Count", "Pipeline Progress", "Month-over-month count of trials completed within the Respective Period", 'Count ("Deal") where Trial Completed Date = Respective Month and Deal Owner = Selected BD Rep Group By Trial Completed Month', "Months in selected period", "Count", "Owner, Month", "Target line = monthly target", "Deal Stage, trial date, Owner", "0", "No", "N/A", "Bar Chart", "Trial completion trend.", "Hover monthly values", "BD Reps", "New", ""),
        # Deals Progressed
        ("BDPERF-9", "Deals Progressed in Stage - Deal Name", "Pipeline Progress", "Name of the deal that progressed to a new stage during the Respective Period", '"Deal" Name (dealname)', "Selected period (stage change date)", "Text", "Deal", "n/a", "dealname", "n/a", "No", "N/A", "Table", "Identify deals with movement.", "Click to open in HubSpot", "BD Reps", "New", ""),
        ("BDPERF-9", "Deals Progressed in Stage - Deal Stage", "Pipeline Progress", "Current deal stage after progression during the Respective Period", '"Deal" Stage (dealstage, decoded)', "Selected period", "Text", "Deal, Stage", "n/a", "dealstage", "n/a", "No", "N/A", "Table", "Stage reached.", "n/a", "BD Reps", "New", ""),
        ("BDPERF-9", "Deals Progressed in Stage - Pipeline", "Pipeline Progress", "Deal pipeline the progressed deal belongs to", '"Deal" Pipeline (pipeline)', "Selected period", "Text", "Pipeline", "n/a", "pipeline", "n/a", "No", "N/A", "Table", "Pipeline context.", "n/a", "BD Reps", "New", ""),
        ("BDPERF-9", "Deals Progressed in Stage - Deal Amount", "Pipeline Progress", "Annual Contract Value of the deal that progressed stage during the Respective Period", '"Deal" Amount (USD) (bookings__usd_)', "Selected period", "Currency, USD", "Deal", "n/a", "bookings__usd_", "USD K", "No", "N/A", "Table", "Deal value.", "n/a", "BD Reps", "New", ""),
        ("BDPERF-9", "Deals Progressed in Stage - Next Steps", "Pipeline Progress", "Next steps recorded on the deal that progressed stage during the Respective Period", '"Deal" Next Step (hs_next_step)', "Selected period", "Text", "Deal", "n/a", "hs_next_step", "n/a", "No", "N/A", "Table", "Follow-up action.", "n/a", "BD Reps", "New", ""),
        # Activities
        ("BDPERF-10", "Activities - Meetings with Existing Clients", "Activities", "Number of meetings held with existing clients during the Respective Period", 'Count ("Meeting") where Associated to Existing Account = True and Meeting Date = Respective Period and Meeting Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Week, Month, Quarter", "n/a", "Meeting date, associations, Owner, hs_meeting_outcome", "0", "No", "N/A", "KPI Number", "Engagement on existing accounts.", "Click for meeting list", "BD Reps", "New", ""),
        ("BDPERF-11", "Activities - Meetings with New Leads", "Activities", "Number of first meetings held with new leads during the Respective Period", 'Count ("Meeting") where Meeting Type = New Business and Meeting Date = Respective Period and Meeting Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Week, Month", "n/a", "Meeting date, Meeting type, Owner", "0", "No", "N/A", "KPI Number", "Top-of-funnel meetings.", "Click for meeting list", "BD Reps", "New", ""),
        ("BDPERF-12", "Activities - Assigned Leads", "Activities", "Number of leads assigned to the rep during the Respective Period", 'Count ("Contact") where Assignment Date = Respective Period and Contact Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Week, Month", "n/a", "Create Date, assignment date, Owner", "0", "No", "N/A", "KPI Number", "Lead assignment volume.", "Click for lead list", "BD Reps", "New", ""),
        ("BDPERF-13", "Activities - Emails Sent", "Activities", "Number of outbound emails sent by the rep during the Respective Period", 'Count ("Engagement") where Engagement Type = EMAIL and Direction = Outbound and Timestamp = Respective Period and Owner = Selected BD Rep', "Selected close period", "Count", "Owner, Week, Month", "n/a", "Engagement type, timestamp, Owner", "0", "No", "N/A", "KPI Number", "Outbound email activity.", "Click for email list", "BD Reps", "New", ""),
    ]

    for m in metrics:
        rows.append({
            "request_id": m[0],
            "metric_name": m[1],
            "tab": m[2],
            "definition": m[3],
            "sql": m[4],
            "window": m[5],
            "fmt": m[6],
            "agg": m[7],
            "threshold": m[8],
            "fields": m[9],
            "std": m[10],
            "filter": m[11],
            "default_filter": m[12],
            "chart": m[13],
            "purpose": m[14],
            "interactions": m[15],
            "audience": m[16],
            "status": m[17],
            "note": m[18] if len(m) > 18 else "",
        })
    return rows


def build_data_source_rows():
    return [
        ("Top Filter", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Close Date", "closedate", "Expected or actual deal close date; drives Close Period filter", "Date", "No", "Internal", "No", "n/a", "2026-06-15", "Quarter / Month / Week / Year grains", "Daily", "RevOps", "Primary time filter"),
        ("Top Filter", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Deal Owner", "hubspot_owner_id", "BD Rep assigned as deal owner", "HubSpot User", "No", "Internal", "No", "Owner.id", "12345678", "Join to owner name dimension", "Daily", "RevOps", "BD Rep filter"),
        ("BDPERF-1", "Goals / Quota", "ANALYTICS_PROD.SALES.QUOTA_TARGET", "Quota", "Quota Target", "quota_target_usd", "Rep quota for period", "Currency", "No", "Internal", "Yes", "Owner + Period", "250000", "Scorecard target source", "Monthly", "Finance", "Confirm source of truth"),
        ("BDPERF-1", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Amount (USD)", "bookings__usd_", "ACV booking amount in USD", "Currency", "No", "Internal", "No", "n/a", "50000", "Core revenue field", "Daily", "RevOps", ""),
        ("BDPERF-1", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Deal Stage", "dealstage", "Current deal stage", "Dropdown", "No", "Internal", "No", "n/a", "closedwon", "Decode stage ID to label", "Daily", "RevOps", ""),
        ("BDPERF-2", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Pipeline", "pipeline", "Deal pipeline name (New Business, Expansion, etc.)", "Dropdown", "No", "Internal", "No", "n/a", "New Business Pipeline", "Won deals breakdown dimension", "Daily", "RevOps", ""),
        ("BDPERF-2", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Close Date", "closedate", "Close date for won deals MoM", "Date", "No", "Internal", "No", "n/a", "", "", "Daily", "RevOps", ""),
        ("BDPERF-3", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Deal Stage Summary", "deal_stage_summary", "Won / Lost / Open summary", "Dropdown", "No", "Internal", "No", "n/a", "Closed Won", "Filter won deals", "Daily", "RevOps", ""),
        ("BDPERF-4", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Forecast Category", "hs_manual_forecast_category", "Commit / Best Case / Pipeline / Omitted", "Dropdown", "No", "Internal", "No", "n/a", "Commit", "Open pipeline table rows", "Daily", "RevOps", ""),
        ("BDPERF-4", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Deal ID", "dealid", "Unique deal identifier for count", "ID", "No", "Internal", "Yes", "n/a", "", "# of deals column", "Daily", "RevOps", ""),
        ("BDPERF-5", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Pipeline Type", "pipeline", "Map to New Business vs Renewal/Expansion for stacked MoM chart", "Dropdown", "No", "Internal", "No", "n/a", "Renewal Pipeline", "Business rule for stack segments", "Daily", "RevOps", "Align stack total with open pipeline table"),
        ("BDPERF-5", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Snapshot Date", "snapshot_date", "Monthly open pipeline snapshot date", "Date", "No", "Internal", "No", "n/a", "2026-06-01", "End-of-month pipeline snapshot", "Daily", "RevOps", ""),
        ("BDPERF-6", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Create Date", "createdate", "Deal creation date for new biz metrics", "Date", "No", "Internal", "No", "n/a", "", "", "Daily", "RevOps", ""),
        ("BDPERF-6", "HubSpot", "ANALYTICS_PROD.SALES.BD_TARGETS", "Target", "New Biz Monthly Target", "new_biz_monthly_target", "Monthly target for new biz deals created", "Count", "No", "Internal", "No", "Owner + Month", "5", "Scaled by period in dashboard", "Monthly", "RevOps", ""),
        ("BDPERF-7", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Demo Completed Date", "demo_completed_date", "Date demo was completed (custom property TBD)", "Date", "No", "Internal", "No", "n/a", "", "Confirm HubSpot property name", "Daily", "RevOps", ""),
        ("BDPERF-7", "HubSpot", "ANALYTICS_PROD.SALES.BD_TARGETS", "Target", "Demo Monthly Target", "demo_monthly_target", "Monthly target for demos completed", "Count", "No", "Internal", "No", "Owner + Month", "2", "", "Monthly", "RevOps", ""),
        ("BDPERF-8", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Trial Completed Date", "trial_completed_date", "Date trial was completed (custom property TBD)", "Date", "No", "Internal", "No", "n/a", "", "Confirm HubSpot property name", "Daily", "RevOps", ""),
        ("BDPERF-8", "HubSpot", "ANALYTICS_PROD.SALES.BD_TARGETS", "Target", "Trial Monthly Target", "trial_monthly_target", "Monthly target for trials completed", "Count", "No", "Internal", "No", "Owner + Month", "1", "", "Monthly", "RevOps", ""),
        ("BDPERF-9", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Deal Name", "dealname", "Deal display name", "Text", "No", "Internal", "Yes", "n/a", "Acme Corp - Enterprise", "", "Daily", "RevOps", ""),
        ("BDPERF-9", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Stage History", "hs_date_entered_*", "Historical stage entry dates to detect progression", "Date", "No", "Internal", "No", "n/a", "", "", "Daily", "RevOps", ""),
        ("BDPERF-9", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_DEAL_SOURCE", "Deal", "Next Step", "hs_next_step", "Next steps on deal", "Text", "No", "Internal", "No", "n/a", "Send revised pricing", "", "Daily", "RevOps", ""),
        ("BDPERF-10", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_ENGAGEMENTS", "Meeting", "Meeting Start Time", "hs_meeting_start_time", "Meeting timestamp", "Date", "No", "Internal", "No", "n/a", "", "", "Daily", "RevOps", ""),
        ("BDPERF-10", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_ENGAGEMENTS", "Meeting", "Associated Company/Deal", "associations", "Link meeting to existing account or open deal", "ID", "No", "Internal", "No", "Company.id, Deal.id", "", "Rule: existing client meeting", "Daily", "RevOps", ""),
        ("BDPERF-11", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_ENGAGEMENTS", "Meeting", "Meeting Type", "hs_activity_type", "New business vs existing classification", "Dropdown", "No", "Internal", "No", "n/a", "New Business", "Define business rules", "Daily", "RevOps", ""),
        ("BDPERF-12", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_CONTACT_SOURCE", "Contact", "Create Date", "createdate", "Lead assignment / create date", "Date", "No", "PII-adjacent", "No", "n/a", "", "", "Daily", "RevOps", ""),
        ("BDPERF-12", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_CONTACT_SOURCE", "Contact", "Contact Owner", "hubspot_owner_id", "Assigned BD Rep", "HubSpot User", "No", "Internal", "No", "Owner.id", "", "", "Daily", "RevOps", ""),
        ("BDPERF-13", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_ENGAGEMENTS", "Engagement", "Engagement Type", "engagement_type", "EMAIL for emails sent", "Dropdown", "No", "Internal", "No", "n/a", "EMAIL", "", "Daily", "RevOps", ""),
        ("BDPERF-13", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_ENGAGEMENTS", "Engagement", "Timestamp", "hs_timestamp", "When email was sent", "Date", "No", "Internal", "No", "n/a", "", "", "Daily", "RevOps", ""),
        ("BDPERF-13", "HubSpot", "ANALYTICS_PROD.HUBSPOT.HUBSPOT_ENGAGEMENTS", "Engagement", "Owner ID", "hubspot_owner_id", "Rep who sent email", "HubSpot User", "No", "Internal", "No", "Owner.id", "", "", "Daily", "RevOps", ""),
    ]


def create_workbook():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # --- 1.2 Scoping ---
    ws12 = wb.create_sheet("1.2_Scoping_Prioritization")
    headers_12 = [
        "Request ID", "Title", "High-Level Objective", "Chart Type (dropdown)",
        "Purpose/Insight", "Owner", "Prioritisation (dropdown)", "Effort (person-weeks)",
        "Strategic Alignment (dropdown)", "Target Audience", "Deadline", "Dependencies/Constraints",
        "Requested By", "Date Requested", "Status", "Notes",
    ]
    for c, h in enumerate(headers_12, 1):
        ws12.cell(1, c, h)
    style_header_row(ws12, len(headers_12))

    today = datetime(2026, 6, 9)
    for i, item in enumerate(build_scoping_rows(), start=2):
        write_row(ws12, i, [
            item["id"], item["title"], item["objective"], item["chart"], item["purpose"],
            "", item["priority"], item["effort"], item["alignment"], item["audience"],
            datetime(2026, 9, 30), "HubSpot deal + engagement data; Quota source TBD",
            "BD Team", today, item["status"],
            "Traffic light: Attainment Green ≥80%, Amber 60–79%, Red <60%; Coverage Green ≥3x, Amber 2–3x, Red <2x; New Biz/Demo/Trial KPI Green ≥ target, Red < target",
        ])

    widths_12 = [18, 32, 45, 30, 50, 28, 13, 23, 32, 27, 26, 38, 19, 25, 28, 60]
    for idx, w in enumerate(widths_12, 1):
        ws12.column_dimensions[get_column_letter(idx)].width = w

    dv_status = DataValidation(type="list", formula1='"Completed,Pending Acceptance,In Progress,New,Blocked"', allow_blank=True)
    ws12.add_data_validation(dv_status)
    dv_status.add(f"O2:O{1 + len(build_scoping_rows())}")

    # --- 1.3 Metrics ---
    ws13 = wb.create_sheet(" 1.3_Metrics (2026 Version)")
    headers_13 = [
        "Request ID", "Metric Name", "Tab #row:columns", "Business Definition (plain English)",
        "SQL/Pseudocode", "Window/Lookback", "Format (e.g., %/count/currency)",
        "Aggregation Types", "Default Threshold/Alert", "Source Fields (comma-separated)",
        "Standartization", "Use as Filter? (Yes/No)", "Default Filter Value",
        "Chart Type (dropdown)", "Purpose/Insight", "Interactions (drilldown, hover, click)",
        "Primary Audience", "Status - Streamlit Apps", "Revops Note - Streamlit Apps ",
    ]
    for c, h in enumerate(headers_13, 1):
        ws13.cell(1, c, h)
    style_header_row(ws13, len(headers_13))

    for i, m in enumerate(build_metrics_rows(), start=2):
        write_row(ws13, i, [
            m["request_id"], m["metric_name"], m["tab"], m["definition"], m["sql"],
            m["window"], m["fmt"], m["agg"], m["threshold"], m["fields"],
            m["std"], m["filter"], m["default_filter"], m["chart"], m["purpose"],
            m["interactions"], m["audience"], m["status"], m.get("note", ""),
        ])

    widths_13 = [14, 28, 22, 45, 40, 18, 16, 22, 22, 30, 12, 14, 16, 22, 40, 35, 20, 18, 30]
    for idx, w in enumerate(widths_13, 1):
        ws13.column_dimensions[get_column_letter(idx)].width = w

    dv_metric_status = DataValidation(
        type="list",
        formula1='"New,In Progress,Pending Acceptance,Need Correction,Completed,Blocked"',
        allow_blank=True,
    )
    ws13.add_data_validation(dv_metric_status)
    dv_metric_status.add(f"R2:R{1 + len(build_metrics_rows())}")

    # --- 1.4 Data Sources ---
    ws14 = wb.create_sheet("1.4_Data_Sources_Fields")
    headers_14 = [
        "Request ID", "Source System (dropdown)", "Database/Schema", "Table / API Endpoint",
        "Field Name", "API Name", "Field Description", "Data Type (dropdown)",
        "Is PII? (dropdown)", "Security Classification (dropdown)", "Primary Key? (Yes/No)",
        "Foreign Key To (table.field)", "Expected Values / Examples", "Transformation Notes",
        "Freshness (e.g., daily 6 AM)", "Owner / SME", "Comments",
    ]
    for c, h in enumerate(headers_14, 1):
        ws14.cell(1, c, h)
    style_header_row(ws14, len(headers_14))

    for i, row in enumerate(build_data_source_rows(), start=2):
        write_row(ws14, i, list(row))

    widths_14 = [14, 16, 38, 22, 24, 28, 32, 14, 10, 18, 12, 20, 24, 30, 18, 14, 28]
    for idx, w in enumerate(widths_14, 1):
        ws14.column_dimensions[get_column_letter(idx)].width = w

    wb.save(OUTPUT_XLSX)
    print(f"Created {OUTPUT_XLSX}")


def create_html_demo():
    html = """<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="UTF-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1.0" />
  <title>BD Performance Dashboard — Demo</title>
  <script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.1/dist/chart.umd.min.js"></script>
  <style>
    :root {
      --bg: #0f1419;
      --surface: #1a2332;
      --surface2: #243044;
      --border: #2d3a4f;
      --text: #e8edf4;
      --muted: #8b9cb3;
      --green: #22c55e;
      --amber: #f59e0b;
      --red: #ef4444;
      --accent: #3b82f6;
    }
    * { box-sizing: border-box; margin: 0; padding: 0; }
    body {
      font-family: "Segoe UI", system-ui, -apple-system, sans-serif;
      background: var(--bg);
      color: var(--text);
      line-height: 1.5;
      min-height: 100vh;
    }
    header {
      background: linear-gradient(135deg, #1e3a5f 0%, #0f1419 100%);
      border-bottom: 1px solid var(--border);
      padding: 1.25rem 2rem;
    }
    header h1 { font-size: 1.5rem; font-weight: 600; }
    header p { color: var(--muted); font-size: 0.875rem; margin-top: 0.25rem; }
    .filters {
      display: flex;
      flex-wrap: wrap;
      gap: 1rem;
      padding: 1rem 2rem;
      background: var(--surface);
      border-bottom: 1px solid var(--border);
      align-items: flex-end;
    }
    .filter-group label {
      display: block;
      font-size: 0.7rem;
      text-transform: uppercase;
      letter-spacing: 0.05em;
      color: var(--muted);
      margin-bottom: 0.35rem;
    }
    .filter-group select {
      background: var(--surface2);
      border: 1px solid var(--border);
      color: var(--text);
      padding: 0.5rem 0.75rem;
      border-radius: 6px;
      font-size: 0.875rem;
      min-width: 160px;
    }
    main { padding: 1.5rem 2rem 3rem; max-width: 1400px; margin: 0 auto; }
    section { margin-bottom: 2.5rem; }
    section h2 {
      font-size: 1rem;
      font-weight: 600;
      margin-bottom: 1rem;
      padding-bottom: 0.5rem;
      border-bottom: 1px solid var(--border);
      display: flex;
      align-items: center;
      gap: 0.5rem;
    }
    section h2 span.num {
      background: var(--accent);
      color: white;
      font-size: 0.7rem;
      padding: 0.15rem 0.5rem;
      border-radius: 4px;
    }
    .legend {
      display: flex;
      gap: 1.25rem;
      flex-wrap: wrap;
      font-size: 0.75rem;
      color: var(--muted);
      margin-bottom: 1.25rem;
    }
    .legend-item { display: flex; align-items: center; gap: 0.4rem; }
    .dot { width: 10px; height: 10px; border-radius: 50%; }
    .dot.green { background: var(--green); }
    .dot.amber { background: var(--amber); }
    .dot.red { background: var(--red); }
    table {
      width: 100%;
      border-collapse: collapse;
      font-size: 0.875rem;
      background: var(--surface);
      border-radius: 8px;
      overflow: hidden;
    }
    th, td {
      padding: 0.75rem 1rem;
      text-align: left;
      border-bottom: 1px solid var(--border);
    }
    th {
      background: var(--surface2);
      font-size: 0.7rem;
      text-transform: uppercase;
      letter-spacing: 0.04em;
      color: var(--muted);
      font-weight: 600;
    }
    tr:last-child td { border-bottom: none; }
    .traffic-light {
      display: inline-flex;
      align-items: center;
      gap: 0.5rem;
      font-weight: 600;
    }
    .light {
      width: 12px;
      height: 12px;
      border-radius: 50%;
      box-shadow: 0 0 8px currentColor;
    }
    .light.green { background: var(--green); color: var(--green); }
    .light.amber { background: var(--amber); color: var(--amber); }
    .light.red { background: var(--red); color: var(--red); }
    .grid-2 { display: grid; grid-template-columns: 1fr 1fr; gap: 1.25rem; }
    .grid-3 { display: grid; grid-template-columns: repeat(3, 1fr); gap: 1rem; }
    @media (max-width: 900px) {
      .grid-2, .grid-3 { grid-template-columns: 1fr; }
    }
    .card {
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 1.25rem;
    }
    .card h3 {
      font-size: 0.8rem;
      color: var(--muted);
      font-weight: 500;
      margin-bottom: 0.75rem;
    }
    .chart-wrap { position: relative; height: 220px; }
    .chart-wrap.tall { height: 260px; }
    .kpi-grid {
      display: grid;
      grid-template-columns: repeat(auto-fill, minmax(200px, 1fr));
      gap: 1rem;
    }
    .kpi {
      background: var(--surface);
      border: 1px solid var(--border);
      border-radius: 10px;
      padding: 1rem 1.25rem;
      border-left: 4px solid var(--border);
    }
    .kpi.green { border-left-color: var(--green); }
    .kpi.amber { border-left-color: var(--amber); }
    .kpi.red { border-left-color: var(--red); }
    .kpi-label { font-size: 0.75rem; color: var(--muted); margin-bottom: 0.25rem; }
    .kpi-value { font-size: 1.75rem; font-weight: 700; }
    .kpi-sub { font-size: 0.75rem; color: var(--muted); margin-top: 0.25rem; }
    .kpi-trend { font-size: 0.8rem; margin-top: 0.35rem; }
    .kpi-trend.up { color: var(--green); }
    .kpi-trend.down { color: var(--red); }
    .kpi-trend.flat { color: var(--muted); }
    .mini-chart { height: 48px; margin-top: 0.5rem; }
    .pipeline-bars { display: flex; flex-direction: column; gap: 0.6rem; }
    .pipe-row {
      display: grid;
      grid-template-columns: 140px 1fr 48px;
      align-items: center;
      gap: 0.75rem;
      font-size: 0.8rem;
    }
    .pipe-bar-bg {
      background: var(--surface2);
      border-radius: 4px;
      height: 22px;
      overflow: hidden;
    }
    .pipe-bar-fill {
      height: 100%;
      border-radius: 4px;
      background: linear-gradient(90deg, var(--accent), #38bdf8);
    }
    .insight-banner {
      background: rgba(245, 158, 11, 0.12);
      border: 1px solid rgba(245, 158, 11, 0.35);
      border-radius: 8px;
      padding: 0.85rem 1rem;
      font-size: 0.85rem;
      margin-bottom: 1.25rem;
      display: flex;
      align-items: flex-start;
      gap: 0.6rem;
    }
    .insight-banner strong { color: var(--amber); }
  </style>
</head>
<body>
  <header>
    <h1>BD Performance Dashboard</h1>
    <p>Weekly performance view for Business Development reps · Q2 2026 · Demo data</p>
  </header>

  <div class="filters">
    <div class="filter-group">
      <label>Close Period</label>
      <select id="closePeriod">
        <option>Close Quarter — Q2 2026</option>
        <option>Close Month — June 2026</option>
        <option>Close Year — FY 2026</option>
        <option>Close Date — Custom</option>
      </select>
    </div>
    <div class="filter-group">
      <label>BD Rep (Owner)</label>
      <select id="bdRep">
        <option selected>Sarah Chen</option>
        <option>James Okonkwo</option>
        <option>Maria Santos</option>
        <option>All Reps (Team View)</option>
      </select>
    </div>
  </div>

  <main>
    <div class="legend">
      <span>Traffic light rules:</span>
      <span class="legend-item"><span class="dot green"></span> Green — on track</span>
      <span class="legend-item"><span class="dot amber"></span> Amber — needs attention</span>
      <span class="legend-item"><span class="dot red"></span> Red — action required</span>
    </div>

    <div class="insight-banner">
      <span>⚠</span>
      <div><strong>Focus areas this week:</strong> Attainment at 72% (amber) — increase pipeline coverage. Outreaches down 12% WoW. Existing deal meetings on track.</div>
    </div>

    <!-- 1. QUOTA ATTAINMENT -->
    <section>
      <h2><span class="num">1</span> Quota Attainment</h2>

      <h3 style="font-size:0.85rem;color:var(--muted);margin-bottom:0.75rem;">1.1 Quota Attainment — Table</h3>
      <table>
        <thead>
          <tr>
            <th>Metric</th>
            <th>Value</th>
            <th>Status</th>
            <th>Notes</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td>Target (Quota)</td>
            <td>$250K</td>
            <td>—</td>
            <td>Q2 2026 assigned quota</td>
          </tr>
          <tr>
            <td>ACV Booked</td>
            <td>$180K</td>
            <td>—</td>
            <td>Closed Won to date</td>
          </tr>
          <tr>
            <td>% Attainment</td>
            <td>72%</td>
            <td><span class="traffic-light"><span class="light amber"></span> Amber</span></td>
            <td>Green ≥80% · Amber 60–79% · Red &lt;60%</td>
          </tr>
          <tr>
            <td>Gap</td>
            <td>$70K</td>
            <td>—</td>
            <td>Remaining to hit quota</td>
          </tr>
          <tr>
            <td>Pipeline Coverage vs Quota</td>
            <td>2.4x</td>
            <td><span class="traffic-light"><span class="light amber"></span> Amber</span></td>
            <td>Green ≥3x · Amber 2–3x · Red &lt;2x · Open pipe $168K / Gap $70K</td>
          </tr>
        </tbody>
      </table>

      <h3 style="font-size:0.85rem;color:var(--muted);margin:1.5rem 0 0.75rem;">1.2 Deals Closed Won This Quarter — by Pipeline</h3>
      <div class="card">
        <div class="pipeline-bars">
          <div class="pipe-row">
            <span>New Business</span>
            <div class="pipe-bar-bg"><div class="pipe-bar-fill" style="width:55%"></div></div>
            <strong>6</strong>
          </div>
          <div class="pipe-row">
            <span>Expansion</span>
            <div class="pipe-bar-bg"><div class="pipe-bar-fill" style="width:36%"></div></div>
            <strong>4</strong>
          </div>
          <div class="pipe-row">
            <span>Transactional</span>
            <div class="pipe-bar-bg"><div class="pipe-bar-fill" style="width:18%"></div></div>
            <strong>2</strong>
          </div>
        </div>
      </div>
    </section>

    <!-- 3. PIPELINE PROGRESS -->
    <section>
      <h2><span class="num">3</span> Pipeline Progress</h2>
      <div class="grid-2">
        <div class="card">
          <h3>Pipeline Amount Closing — Won vs Lost</h3>
          <div class="chart-wrap"><canvas id="pipeAmountPie"></canvas></div>
        </div>
        <div class="card">
          <h3>Pipeline Amount Closing — MoM Trend</h3>
          <div class="chart-wrap"><canvas id="pipeAmountTrend"></canvas></div>
        </div>
        <div class="card">
          <h3>Deals Closing — Won vs Lost</h3>
          <div class="chart-wrap"><canvas id="dealsPie"></canvas></div>
        </div>
        <div class="card">
          <h3>Deals Closing — MoM Trend</h3>
          <div class="chart-wrap"><canvas id="dealsTrend"></canvas></div>
        </div>
      </div>

      <h3 style="font-size:0.85rem;color:var(--muted);margin:1.5rem 0 0.75rem;">Pipeline Coverage vs Quota — by Forecast Category</h3>
      <table>
        <thead>
          <tr>
            <th>Forecast Category</th>
            <th>Pipeline ACV</th>
            <th>Coverage vs Gap</th>
            <th>Status</th>
          </tr>
        </thead>
        <tbody>
          <tr>
            <td>Commit</td>
            <td>$85K</td>
            <td>1.2x</td>
            <td><span class="traffic-light"><span class="light green"></span> Green</span></td>
          </tr>
          <tr>
            <td>Best Case</td>
            <td>$52K</td>
            <td>0.7x</td>
            <td><span class="traffic-light"><span class="light amber"></span> Amber</span></td>
          </tr>
          <tr>
            <td>Pipeline</td>
            <td>$31K</td>
            <td>0.4x</td>
            <td><span class="traffic-light"><span class="light red"></span> Red</span></td>
          </tr>
          <tr>
            <td><strong>Total Open Pipeline</strong></td>
            <td><strong>$168K</strong></td>
            <td><strong>2.4x</strong></td>
            <td><span class="traffic-light"><span class="light amber"></span> Amber</span></td>
          </tr>
        </tbody>
      </table>

      <div class="grid-2" style="margin-top:1.25rem;">
        <div class="card">
          <h3>New Deals Created — Amount (bars) + Count (line) MoM</h3>
          <div class="chart-wrap tall"><canvas id="newDealsCombo"></canvas></div>
        </div>
        <div class="card">
          <h3>Deals Progressed in Stage</h3>
          <table style="font-size:0.8rem;">
            <thead>
              <tr>
                <th>Deal</th>
                <th>Stage</th>
                <th>Amount</th>
                <th>Days in Stage</th>
              </tr>
            </thead>
            <tbody>
              <tr><td>Acme Corp — Enterprise</td><td>Proposal</td><td>$45K</td><td><span class="traffic-light"><span class="light green"></span> 8d</span></td></tr>
              <tr><td>GlobalTech Ltd</td><td>Negotiation</td><td>$62K</td><td><span class="traffic-light"><span class="light amber"></span> 24d</span></td></tr>
              <tr><td>EduFirst Academy</td><td>Discovery</td><td>$28K</td><td><span class="traffic-light"><span class="light green"></span> 5d</span></td></tr>
              <tr><td>MedHealth Group</td><td>Demo Completed</td><td>$38K</td><td><span class="traffic-light"><span class="light red"></span> 52d</span></td></tr>
              <tr><td>FinServe APAC</td><td>Contract Sent</td><td>$55K</td><td><span class="traffic-light"><span class="light amber"></span> 18d</span></td></tr>
            </tbody>
          </table>
        </div>
      </div>
    </section>

    <!-- 4. ACTIVITIES -->
    <section>
      <h2><span class="num">4</span> Activities</h2>
      <p style="font-size:0.8rem;color:var(--muted);margin-bottom:1rem;">Each metric shows period total + WoW trend (lookback scales with selected close period)</p>
      <div class="kpi-grid">
        <div class="kpi green">
          <div class="kpi-label">Existing Deal Meetings</div>
          <div class="kpi-value">18</div>
          <div class="kpi-sub">Selected period (Q2 to date)</div>
          <div class="kpi-trend up">▲ +8% WoW</div>
          <canvas class="mini-chart" id="kpi1"></canvas>
        </div>
        <div class="kpi amber">
          <div class="kpi-label">New Deal Meetings</div>
          <div class="kpi-value">7</div>
          <div class="kpi-sub">Selected period</div>
          <div class="kpi-trend down">▼ −14% WoW</div>
          <canvas class="mini-chart" id="kpi2"></canvas>
        </div>
        <div class="kpi green">
          <div class="kpi-label">Inbound Marketing Leads</div>
          <div class="kpi-value">24</div>
          <div class="kpi-sub">Selected period</div>
          <div class="kpi-trend up">▲ +5% WoW</div>
          <canvas class="mini-chart" id="kpi3"></canvas>
        </div>
        <div class="kpi red">
          <div class="kpi-label">Leads Disqualified</div>
          <div class="kpi-value">9</div>
          <div class="kpi-sub">Selected period · 38% of inbound</div>
          <div class="kpi-trend up">▲ +22% WoW</div>
          <canvas class="mini-chart" id="kpi4"></canvas>
        </div>
        <div class="kpi amber">
          <div class="kpi-label">Outreaches</div>
          <div class="kpi-value">142</div>
          <div class="kpi-sub">Selected period</div>
          <div class="kpi-trend down">▼ −12% WoW</div>
          <canvas class="mini-chart" id="kpi5"></canvas>
        </div>
      </div>
    </section>
  </main>

  <script>
    Chart.defaults.color = "#8b9cb3";
    Chart.defaults.borderColor = "#2d3a4f";
    const months = ["Jan", "Feb", "Mar", "Apr", "May", "Jun"];

    new Chart(document.getElementById("pipeAmountPie"), {
      type: "doughnut",
      data: {
        labels: ["Won ($180K)", "Lost ($95K)"],
        datasets: [{ data: [180, 95], backgroundColor: ["#22c55e", "#ef4444"], borderWidth: 0 }]
      },
      options: { plugins: { legend: { position: "bottom" } }, maintainAspectRatio: false }
    });

    new Chart(document.getElementById("pipeAmountTrend"), {
      type: "line",
      data: {
        labels: months,
        datasets: [
          { label: "Won ACV ($K)", data: [22, 28, 35, 30, 38, 27], borderColor: "#22c55e", tension: 0.3, fill: false },
          { label: "Lost ACV ($K)", data: [12, 18, 14, 20, 16, 15], borderColor: "#ef4444", tension: 0.3, fill: false }
        ]
      },
      options: { plugins: { legend: { position: "bottom" } }, scales: { y: { beginAtZero: true } }, maintainAspectRatio: false }
    });

    new Chart(document.getElementById("dealsPie"), {
      type: "doughnut",
      data: {
        labels: ["Won (12 deals)", "Lost (8 deals)"],
        datasets: [{ data: [12, 8], backgroundColor: ["#22c55e", "#ef4444"], borderWidth: 0 }]
      },
      options: { plugins: { legend: { position: "bottom" } }, maintainAspectRatio: false }
    });

    new Chart(document.getElementById("dealsTrend"), {
      type: "line",
      data: {
        labels: months,
        datasets: [
          { label: "Won", data: [1, 2, 3, 2, 2, 2], borderColor: "#22c55e", tension: 0.3 },
          { label: "Lost", data: [1, 1, 2, 1, 2, 1], borderColor: "#ef4444", tension: 0.3 }
        ]
      },
      options: { plugins: { legend: { position: "bottom" } }, scales: { y: { beginAtZero: true, ticks: { stepSize: 1 } } }, maintainAspectRatio: false }
    });

    new Chart(document.getElementById("newDealsCombo"), {
      type: "bar",
      data: {
        labels: months,
        datasets: [
          { label: "New Deal ACV ($K)", data: [45, 52, 68, 55, 72, 48], backgroundColor: "rgba(59,130,246,0.7)", yAxisID: "y" },
          { label: "# New Deals", data: [3, 4, 5, 4, 6, 4], type: "line", borderColor: "#f59e0b", tension: 0.3, yAxisID: "y1" }
        ]
      },
      options: {
        plugins: { legend: { position: "bottom" } },
        scales: {
          y: { beginAtZero: true, position: "left", title: { display: true, text: "ACV ($K)" } },
          y1: { beginAtZero: true, position: "right", grid: { drawOnChartArea: false }, title: { display: true, text: "Count" } }
        },
        maintainAspectRatio: false
      }
    });

    function miniSpark(id, data, color) {
      new Chart(document.getElementById(id), {
        type: "line",
        data: { labels: ["W1","W2","W3","W4","W5","W6","W7","W8"], datasets: [{ data, borderColor: color, borderWidth: 2, pointRadius: 0, tension: 0.35, fill: false }] },
        options: { plugins: { legend: { display: false } }, scales: { x: { display: false }, y: { display: false } }, maintainAspectRatio: false }
      });
    }
    miniSpark("kpi1", [2,3,2,4,3,4,5,5], "#22c55e");
    miniSpark("kpi2", [2,1,2,1,2,1,1,1], "#f59e0b");
    miniSpark("kpi3", [2,3,3,2,3,3,4,4], "#22c55e");
    miniSpark("kpi4", [1,1,1,2,1,2,2,3], "#ef4444");
    miniSpark("kpi5", [22,20,21,19,18,20,17,16], "#f59e0b");
  </script>
</body>
</html>"""

    with open(OUTPUT_HTML, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"Created {OUTPUT_HTML}")


if __name__ == "__main__":
    create_workbook()
